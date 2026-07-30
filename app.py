from flask import Flask, render_template, request, redirect
import sqlite3
import requests
from dotenv import load_dotenv
import os
from functools import wraps
from flask import session, Response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'dev-fallback-key')
DATABASE = os.getenv('DATABASE', 'running.db')
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', '')

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(f'/login?next=/{f.__name__}')
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form['password'] == ADMIN_PASSWORD:
            session['logged_in'] = True
            return redirect('/')
        error = 'Wrong password'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def fetch_weather(date):
    url = "https://archive-api.open-meteo.com/v1/archive"
    params = {
        "latitude": 52.23,
        "longitude": 21.01,
        "start_date": date,
        "end_date": date,
        "daily": "temperature_2m_max",
        "hourly": "relativehumidity_2m,surface_pressure",
        "timezone": "Europe/Warsaw"
    }
    try:
        r = requests.get(url, params=params, timeout=5)
        d = r.json()
        temperature = d["daily"]["temperature_2m_max"][0]
        humidity = d["hourly"]["relativehumidity_2m"][12]
        pressure = round(d["hourly"]["surface_pressure"][12])
        return temperature, pressure, humidity
    except:
        return None, None, None


def fetch_prev_pressure(date):
    """Pobiera ciśnienie z dnia poprzedniego"""
    from datetime import datetime, timedelta
    prev_date = (datetime.strptime(date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')

    url = "https://archive-api.open-meteo.com/v1/archive"
    params = {
        "latitude": 51.1,
        "longitude": 17.03,
        "start_date": prev_date,
        "end_date": prev_date,
        "hourly": "surface_pressure",
        "timezone": "Europe/Warsaw"
    }
    try:
        r = requests.get(url, params=params, timeout=5)
        d = r.json()
        return round(d["hourly"]["surface_pressure"][12])
    except:
        return None

@app.route('/')
def index():
    db = get_db()
    runs = db.execute("""
        SELECT r.*, w.headache, w.energy_before, w.energy_after, w.notes,
               p.temperature, p.pressure, p.humidity
        FROM runs r
        LEFT JOIN wellbeing w ON w.run_id = r.id
        LEFT JOIN weather p ON p.date = r.date
        ORDER BY r.date DESC
    """).fetchall()
    return render_template('index.html', runs=runs)

@app.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    if request.method == 'POST':
        db = get_db()

        db.execute("""
            INSERT INTO runs (date, distance_km, duration_min, avg_pace, avg_speed,
                calories, avg_heart_rate, max_heart_rate, cadence,
                training_effect, vo2max, recovery_time_h)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            request.form['date'],
            request.form['distance_km'],
            request.form['duration_min'],
            request.form['avg_pace'],
            request.form['avg_speed'],
            request.form['calories'],
            request.form['avg_heart_rate'],
            request.form['max_heart_rate'],
            request.form['cadence'],
            request.form['training_effect'],
            request.form['vo2max'],
            request.form['recovery_time_h'],
        ))
        run_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        db.execute("""
            INSERT INTO wellbeing (run_id, headache, energy_before, energy_after, notes)
            VALUES (?,?,?,?,?)
        """, (
            run_id,
            request.form['headache'],
            request.form['energy_before'],
            request.form['energy_after'],
            request.form.get('notes', '')
        ))

        date = request.form['date']
        temperature, pressure, humidity = fetch_weather(date)
        if temperature is not None:
            db.execute("""
                INSERT OR IGNORE INTO weather (date, temperature, pressure, humidity)
                VALUES (?,?,?,?)
            """, (date, temperature, pressure, humidity))

        pressure_prev = fetch_prev_pressure(date)
        if pressure_prev is not None:
            db.execute("""
                INSERT OR IGNORE INTO weather_prev (date, pressure_prev)
                VALUES (?, ?)
            """, (date, pressure_prev))

        db.commit()
        return redirect('/')

    return render_template('add.html')


@app.route('/stats')
def stats():
    db = get_db()

    summary = db.execute("""
        SELECT 
            COUNT(*) as total_runs,
            ROUND(SUM(distance_km), 1) as total_distance,
            ROUND(AVG(distance_km), 1) as avg_distance,
            ROUND(AVG(avg_heart_rate), 0) as avg_hr,
            ROUND(AVG(vo2max), 1) as avg_vo2max,
            ROUND(AVG(calories), 0) as avg_calories
        FROM runs
    """).fetchone()

    runs_over_time = db.execute("""
        SELECT r.date, r.distance_km, r.avg_heart_rate, r.avg_speed, r.vo2max,
               r.recovery_time_h, r.calories, r.training_effect,
               w.pressure, wp.pressure_prev,
               w.pressure - wp.pressure_prev as pressure_change
        FROM runs r
        LEFT JOIN weather w ON w.date = r.date
        LEFT JOIN weather_prev wp ON wp.date = r.date
        ORDER BY r.date ASC
    """).fetchall()

    weather_vs_hr = db.execute("""
        SELECT r.date, r.avg_heart_rate, w.temperature, w.humidity,
               s.headache
        FROM runs r
        LEFT JOIN weather w ON w.date = r.date
        LEFT JOIN wellbeing s ON s.run_id = r.id
        ORDER BY r.date ASC
    """).fetchall()

    headache_stats = db.execute("""
        SELECT 
            s.headache,
            COUNT(*) as count,
            ROUND(AVG(w.temperature), 1) as avg_temp,
            ROUND(AVG(w.humidity), 1) as avg_humidity,
            ROUND(AVG(r.max_heart_rate), 0) as avg_hr,
            ROUND(AVG(r.training_effect), 1) as training_effect
        FROM wellbeing s
        JOIN runs r ON r.id = s.run_id
        LEFT JOIN weather w ON w.date = r.date
        GROUP BY s.headache
    """).fetchall()

    wellbeing_over_time = db.execute("""
        SELECT r.date, s.energy_before, s.energy_after, s.headache, s.notes
        FROM runs r
        JOIN wellbeing s ON s.run_id = r.id
        ORDER BY r.date ASC
    """).fetchall()

    return render_template('stats.html',
                           summary=summary,
                           runs_over_time=[dict(r) for r in runs_over_time],
                           weather_vs_hr=[dict(r) for r in weather_vs_hr],
                           headache_stats=[dict(r) for r in headache_stats],
                           wellbeing_over_time=[dict(r) for r in wellbeing_over_time]
                           )

@app.route('/rides')
def rides():
    db = get_db()
    rides = db.execute("""
        SELECT r.*, w.headache, w.energy_before, w.energy_after, w.notes,
               p.temperature, p.pressure, p.humidity
        FROM rides r
        LEFT JOIN wellbeing_rides w ON w.ride_id = r.id
        LEFT JOIN weather p ON p.date = r.date
        ORDER BY r.date DESC
    """).fetchall()
    return render_template('rides.html', rides=rides)

@app.route('/add_ride', methods=['GET', 'POST'])
@login_required
def add_ride():
    if request.method == 'POST':
        db = get_db()

        db.execute("""
            INSERT INTO rides (date, distance_km, duration_min, avg_speed, max_speed,
                calories, avg_heart_rate, max_heart_rate, training_effect, recovery_time_h)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            request.form['date'],
            request.form['distance_km'],
            request.form['duration_min'],
            request.form['avg_speed'],
            request.form['max_speed'],
            request.form['calories'],
            request.form['avg_heart_rate'],
            request.form['max_heart_rate'],
            request.form['training_effect'],
            request.form['recovery_time_h'],
        ))
        ride_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        db.execute("""
            INSERT INTO wellbeing_rides (ride_id, headache, energy_before, energy_after, notes)
            VALUES (?,?,?,?,?)
        """, (
            ride_id,
            request.form['headache'],
            request.form['energy_before'],
            request.form['energy_after'],
            request.form.get('notes', '')
        ))

        date = request.form['date']
        temperature, pressure, humidity = fetch_weather(date)
        if temperature is not None:
            db.execute("""
                INSERT OR IGNORE INTO weather (date, temperature, pressure, humidity)
                VALUES (?,?,?,?)
            """, (date, temperature, pressure, humidity))

        pressure_prev = fetch_prev_pressure(date)
        if pressure_prev is not None:
            db.execute("""
                INSERT OR IGNORE INTO weather_prev (date, pressure_prev)
                VALUES (?, ?)
            """, (date, pressure_prev))

        db.commit()
        return redirect('/rides')

    return render_template('add_ride.html')

@app.route('/stats_rides')
def stats_rides():
    db = get_db()

    summary = db.execute("""
        SELECT
            COUNT(*) as total_rides,
            ROUND(SUM(distance_km), 1) as total_distance,
            ROUND(AVG(distance_km), 1) as avg_distance,
            ROUND(AVG(avg_heart_rate), 0) as avg_hr,
            ROUND(AVG(training_effect), 1) as avg_training_effect,
            ROUND(AVG(calories), 0) as avg_calories
        FROM rides
    """).fetchone()

    rides_over_time = db.execute("""
        SELECT r.date, r.distance_km, r.avg_heart_rate, r.avg_speed,
               r.max_speed, r.calories, r.training_effect, r.recovery_time_h,
               w.pressure, wp.pressure_prev,
               w.pressure - wp.pressure_prev as pressure_change
        FROM rides r
        LEFT JOIN weather w ON w.date = r.date
        LEFT JOIN weather_prev wp ON wp.date = r.date
        ORDER BY r.date ASC
    """).fetchall()

    weather_vs_hr = db.execute("""
        SELECT r.date, r.avg_heart_rate, w.temperature, w.humidity,
               wr.headache
        FROM rides r
        LEFT JOIN weather w ON w.date = r.date
        LEFT JOIN wellbeing_rides wr ON wr.ride_id = r.id
        ORDER BY r.date ASC
    """).fetchall()

    headache_stats = db.execute("""
        SELECT
            wr.headache,
            COUNT(*) as count,
            ROUND(AVG(w.temperature), 1) as avg_temp,
            ROUND(AVG(w.humidity), 1) as avg_humidity,
            ROUND(AVG(r.max_heart_rate), 0) as avg_hr,
            ROUND(AVG(r.training_effect), 1) as avg_training_effect
        FROM wellbeing_rides wr
        JOIN rides r ON r.id = wr.ride_id
        LEFT JOIN weather w ON w.date = r.date
        GROUP BY wr.headache
    """).fetchall()

    wellbeing_over_time = db.execute("""
        SELECT r.date, wr.energy_before, wr.energy_after, wr.headache, wr.notes
        FROM rides r
        JOIN wellbeing_rides wr ON wr.ride_id = r.id
        ORDER BY r.date ASC
    """).fetchall()

    return render_template('stats_rides.html',
        summary=summary,
        rides_over_time=[dict(r) for r in rides_over_time],
        weather_vs_hr=[dict(r) for r in weather_vs_hr],
        headache_stats=[dict(r) for r in headache_stats],
        wellbeing_over_time=[dict(r) for r in wellbeing_over_time]
    )
@app.route('/compare')
def compare():
    db = get_db()

    runs_summary = db.execute("""
        SELECT 
            COUNT(*) as total,
            ROUND(SUM(distance_km), 1) as total_distance,
            ROUND(AVG(calories), 0) as avg_calories,
            ROUND(SUM(calories), 0) as total_calories,
            ROUND(AVG(avg_heart_rate), 0) as avg_hr,
            ROUND(AVG(recovery_time_h), 1) as avg_recovery
        FROM runs
    """).fetchone()

    rides_summary = db.execute("""
        SELECT 
            COUNT(*) as total,
            ROUND(SUM(distance_km), 1) as total_distance,
            ROUND(AVG(calories), 0) as avg_calories,
            ROUND(SUM(calories), 0) as total_calories,
            ROUND(AVG(avg_heart_rate), 0) as avg_hr,
            ROUND(AVG(recovery_time_h), 1) as avg_recovery
        FROM rides
    """).fetchone()

    runs_over_time = db.execute("""
        SELECT date, calories, avg_heart_rate, recovery_time_h
        FROM runs ORDER BY date ASC
    """).fetchall()

    rides_over_time = db.execute("""
        SELECT date, calories, avg_heart_rate, recovery_time_h
        FROM rides ORDER BY date ASC
    """).fetchall()

    return render_template('compare.html',
        runs_summary=dict(runs_summary),
        rides_summary=dict(rides_summary),
        runs_over_time=[dict(r) for r in runs_over_time],
        rides_over_time=[dict(r) for r in rides_over_time]
    )
def make_xlsx(headers, rows, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="E05C00")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route('/export/runs')
def export_runs():
    db = get_db()
    runs = db.execute("""
        SELECT r.date, r.distance_km, r.duration_min, r.avg_pace, r.avg_speed,
               r.calories, r.avg_heart_rate, r.max_heart_rate, r.cadence,
               r.training_effect, r.vo2max, r.recovery_time_h,
               w.temperature, w.pressure, w.humidity,
               w.pressure - wp.pressure_prev as pressure_change,
               s.headache, s.energy_before, s.energy_after, s.notes
        FROM runs r
        LEFT JOIN weather w ON w.date = r.date
        LEFT JOIN weather_prev wp ON wp.date = r.date
        LEFT JOIN wellbeing s ON s.run_id = r.id
        ORDER BY r.date ASC
    """).fetchall()

    headers = [
        'Date', 'Distance (km)', 'Duration (min)', 'Avg Pace', 'Avg Speed (km/h)',
        'Calories', 'Avg HR', 'Max HR', 'Cadence',
        'Training Effect', 'VO2Max', 'Recovery (h)',
        'Temperature (°C)', 'Pressure (hPa)', 'Humidity (%)', 'Pressure Change (hPa)',
        'Headache', 'Energy Before', 'Energy After', 'Notes'
    ]

    output = make_xlsx(headers, runs, 'Runs')
    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=runs.xlsx'}
    )

@app.route('/export/rides')
def export_rides():
    db = get_db()
    rides = db.execute("""
        SELECT r.date, r.distance_km, r.duration_min, r.avg_speed, r.max_speed,
               r.calories, r.avg_heart_rate, r.max_heart_rate,
               r.training_effect, r.recovery_time_h,
               w.temperature, w.pressure, w.humidity,
               w.pressure - wp.pressure_prev as pressure_change,
               s.headache, s.energy_before, s.energy_after, s.notes
        FROM rides r
        LEFT JOIN weather w ON w.date = r.date
        LEFT JOIN weather_prev wp ON wp.date = r.date
        LEFT JOIN wellbeing_rides s ON s.ride_id = r.id
        ORDER BY r.date ASC
    """).fetchall()

    headers = [
        'Date', 'Distance (km)', 'Duration (min)', 'Avg Speed (km/h)', 'Max Speed (km/h)',
        'Calories', 'Avg HR', 'Max HR',
        'Training Effect', 'Recovery (h)',
        'Temperature (°C)', 'Pressure (hPa)', 'Humidity (%)', 'Pressure Change (hPa)',
        'Headache', 'Energy Before', 'Energy After', 'Notes'
    ]

    output = make_xlsx(headers, rides, 'Rides')
    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=rides.xlsx'}
    )

@app.route('/export/all')
def export_all():
    db = get_db()

    runs = db.execute("""
        SELECT 'Run', r.date, r.distance_km, r.duration_min, r.avg_speed,
               r.calories, r.avg_heart_rate, r.max_heart_rate,
               r.training_effect, r.recovery_time_h,
               w.temperature, w.pressure, w.humidity,
               w.pressure - wp.pressure_prev,
               s.headache, s.energy_before, s.energy_after, s.notes
        FROM runs r
        LEFT JOIN weather w ON w.date = r.date
        LEFT JOIN weather_prev wp ON wp.date = r.date
        LEFT JOIN wellbeing s ON s.run_id = r.id
    """).fetchall()

    rides = db.execute("""
        SELECT 'Ride', r.date, r.distance_km, r.duration_min, r.avg_speed,
               r.calories, r.avg_heart_rate, r.max_heart_rate,
               r.training_effect, r.recovery_time_h,
               w.temperature, w.pressure, w.humidity,
               w.pressure - wp.pressure_prev,
               s.headache, s.energy_before, s.energy_after, s.notes
        FROM rides r
        LEFT JOIN weather w ON w.date = r.date
        LEFT JOIN weather_prev wp ON wp.date = r.date
        LEFT JOIN wellbeing_rides s ON s.ride_id = r.id
    """).fetchall()

    all_activities = sorted(list(runs) + list(rides), key=lambda x: x[1])

    headers = [
        'Type', 'Date', 'Distance (km)', 'Duration (min)', 'Avg Speed (km/h)',
        'Calories', 'Avg HR', 'Max HR',
        'Training Effect', 'Recovery (h)',
        'Temperature (°C)', 'Pressure (hPa)', 'Humidity (%)', 'Pressure Change (hPa)',
        'Headache', 'Energy Before', 'Energy After', 'Notes'
    ]

    output = make_xlsx(headers, all_activities, 'All Activities')
    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=activities.xlsx'}
    )

if __name__ == '__main__':
    app.run(debug=True)